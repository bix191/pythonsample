#!/bin/env python3
# -*- coding:utf-8 -*

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl import utils
import sqlite3
import datetime
from cellutils import getMergedCellValue
import calendar
from openpyxl.styles.borders import Border,Side

params={
    "sheet":"2019下期工数",
    "staff":"Aさん",
    "startyear":2020,
    "startmonth":3
}


wb = load_workbook("data.xlsx")
ws = wb[params["sheet"]]
staff=params["staff"]
monthidx=(params["startmonth"]-4)%6
print(monthidx)
if monthidx<0:
    monthidx=-monthidx
    monthidx=monthidx+5

print(monthidx)
workcodes={}

for row in ws.iter_rows(min_row=2):
    cell_person=getMergedCellValue(ws,row[0])
    cell_workcode=getMergedCellValue(ws,row[1])
    if staff==cell_person:
        manhour=row[2+monthidx].value
        workcodes[cell_workcode]= { 
            "person": cell_person,
            "manhour": manhour
        }
workcodes["no-code"] = {
    "person": cell_person,
    "manhour": 0
}

ws2=wb.create_sheet(staff+"今月作業")
ws2.column_dimensions[utils.get_column_letter(1)].width=10

headrow=["日付","日合計"]
for workcode in workcodes.keys():
    headrow.append(workcode)

ws2.append(headrow)


startdate=datetime.datetime(params["startyear"],params["startmonth"],1,0,0,0,0)

daycnt=calendar.monthrange(params["startyear"],params["startmonth"])[1]

for idx,date1 in enumerate(startdate+datetime.timedelta(n) for n in range(daycnt)):
    datarow=[]
    datarow.append(date1.date())

    datesumstr="=SUM("+utils.get_column_letter(3)+str(idx+2)+":"+utils.get_column_letter(3+len(workcodes))+str(idx+2)+")"

    datarow.append(datesumstr)
    for idx in range(len(workcodes)):
        datarow.append(0)
              
    ws2.append(datarow)          


sumstrs=["合計"]
for idx in range(len(workcodes)+1):
    sumstr="=SUM("+utils.get_column_letter(2+idx)+"2:"+utils.get_column_letter(2+idx)+str(2+daycnt-1)+")"
    sumstrs.append(sumstr)
    
ws2.append(sumstrs)


plansumstrs=["予定合計",""]
for key in workcodes.keys():
    plansumstr=workcodes[key]["manhour"]
    plansumstrs.append(plansumstr)


ws2.append(plansumstrs)


#罫線
side=Side(style="thin",color="000000")
border=Border(top=side,bottom=side,left=side,right=side)
for row in ws2:
    for cell in row:
        ws2[cell.coordinate].border=border
    

wb.save("data2.xlsx")

